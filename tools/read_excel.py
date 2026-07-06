"""Dump all sheets from the strategy Excel file to readable text."""
import openpyxl, sys

path = r"List selected and to be selected.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for shname in wb.sheetnames:
    ws = wb[shname]
    print(f"\n{'='*80}")
    print(f"SHEET: {shname}  ({ws.max_row} rows × {ws.max_column} cols)")
    print('='*80)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        # Skip entirely empty rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        cells = [str(v).strip() if v is not None else '' for v in row]
        print(' | '.join(c for c in cells))
